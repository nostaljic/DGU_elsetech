import glob
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import datetime
import shutil
import os
from .geocoding import *

from get_water_info.models import AnalysisData
from login.models import Member
from water_req.models import AnalysisRequest

NEW_FILE_PATH = os.path.join(os.getcwd(), "get_water_info\\newData", '*.xlsx')
OLD_FILE_PATH = os.path.join(os.getcwd(), "get_water_info\\oldData")


# 엑셀 파일 읽어서 DB 저장
def read_new_data(path, new_path):
    filename = os.path.basename(path)
    old_data_path = os.path.join(new_path, filename)

    wb = openpyxl.load_workbook(path)
    ws = wb['Sheet1']

    member_id = ws['C7'].value   # 회원 아이디
    member_id = Member.objects.get(pk=member_id)

    request_date = ws['A6'].value  # 요청 날짜
    request_date = request_date.strftime('%Y-%m-%d')
    request_date = AnalysisRequest.objects.get(pk=request_date)

    location = ws['C10'].value
    crd = get_geocode(location)
    latitude = crd[0]
    longitude = crd[1]

    name = ws['A7'].value  # 이름
    water_origin = ws['C18'].value  # 취수원
    turbidity = ws['E18'].value  # 탁도
    fe_origin = ws['H18'].value  # 철

    cell_row = ord('D')
    cell_pos = chr(cell_row) + str(21)
    
    # analysis_data 순서
    # label = ['철(Fe)', '망간(Mn)', '알루미눔(Al)', '육안검사', '종합평가']
    analysis_data = []

    while ws[cell_pos].value:
        day = ws[cell_pos].value
        day = int(day[:-2])

        for i in range(5):
            # 이미지 읽기
            pos = chr(cell_row) + str(i + 22)
            if i == 3:
                # image_loader = SheetImageLoader(ws)
                # image = image_loader.get(pos)
                analysis_data.append('이미지 URL')

                continue

            value = ws[pos].value
            analysis_data.append(value)

        # DB에 저장
        data = AnalysisData(
            member_id=member_id,
            request_date=request_date,
            name=name,
            filename=filename,
            location=location,
            latitude=latitude,
            longitude=longitude,
            water_origin=water_origin,
            fe_origin=fe_origin,
            turbidity=turbidity,
            date=day,
            fe_user=analysis_data[0],
            mn_user=analysis_data[1],
            al_user=analysis_data[2],
            img=analysis_data[3],
            total=analysis_data[4],
        )
        data.save()

        cell_row += 1
        cell_pos = chr(cell_row) + str(21)

    # 읽은 파일 이동
    shutil.move(path, old_data_path)


def read_all_data(test=False):
    global NEW_FILE_PATH, OLD_FILE_PATH

    if test:
        NEW_FILE_PATH = os.path.join(os.getcwd(), "newData", '*.xlsx')
        OLD_FILE_PATH = os.path.join(os.getcwd(), "oldData")

    path = NEW_FILE_PATH
    new_path = OLD_FILE_PATH

    file_list = glob.glob(path)
    for file in file_list:
        read_new_data(file, new_path)


# test
'''
if __name__ == '__main__':
    read_all_data(test=True)
'''
